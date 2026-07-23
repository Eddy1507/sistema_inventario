from flask import Blueprint, render_template, request, send_file
from sqlalchemy import and_
from datetime import datetime, timedelta
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

from models.database import (
    Producto,
    Cliente,
    Proveedor,
    Venta,
    Compra
)


reportes_bp = Blueprint("reportes", __name__)

@reportes_bp.route("/reportes")
def reportes():

    fecha_inicio = request.args.get("inicio")
    fecha_fin = request.args.get("fin")

    ventas = Venta.query

    if fecha_inicio and fecha_fin:

        inicio = datetime.strptime(
            fecha_inicio,
            "%Y-%m-%d"
        )

        fin = datetime.strptime(
            fecha_fin,
            "%Y-%m-%d"
        ) + timedelta(days=1)

        ventas = ventas.filter(
            Venta.fecha >= inicio,
            Venta.fecha < fin
        )

    ventas = ventas.order_by(
        Venta.fecha.desc()
    ).all()

    total_productos = Producto.query.count()
    total_clientes = Cliente.query.count()
    total_proveedores = Proveedor.query.count()
    total_ventas = Venta.query.count()
    total_compras = Compra.query.count()

    productos_stock_bajo = Producto.query.filter(
        Producto.stock <= 10
    ).all()

    valor_inventario = sum(
        p.precio * p.stock
        for p in Producto.query.all()
    )

    total_periodo = sum(
        venta.total
        for venta in ventas
    )

    return render_template(
        "reportes.html",
        ventas=ventas,
        total_periodo=total_periodo,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        total_productos=total_productos,
        total_clientes=total_clientes,
        total_proveedores=total_proveedores,
        total_ventas=total_ventas,
        total_compras=total_compras,
        valor_inventario=valor_inventario,
        productos_stock_bajo=productos_stock_bajo
    )

@reportes_bp.route("/reportes/pdf")
def reporte_pdf():

    ventas = Venta.query.order_by(
        Venta.fecha.desc()
    ).all()

    buffer = io.BytesIO()

    pdf = SimpleDocTemplate(
        buffer,
        pagesize=(21*cm, 29.7*cm)
    )

    estilos = getSampleStyleSheet()

    elementos = []

    elementos.append(
        Paragraph("<b>EMYE INVENTORY</b>", estilos["Title"])
    )

    elementos.append(
        Paragraph("Reporte de Ventas", estilos["Heading2"])
    )

    datos = [
        ["ID", "Fecha", "Cliente", "Total"]
    ]

    total = 0

    for venta in ventas:

        datos.append([

            str(venta.id),

            venta.fecha.strftime("%d/%m/%Y"),

            venta.cliente.nombre,

            f"${venta.total:.2f}"

        ])

        total += venta.total

    datos.append([
        "",
        "",
        "TOTAL",
        f"${total:.2f}"
    ])

    tabla = Table(datos)

    tabla.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.darkblue),

        ("TEXTCOLOR",(0,0),(-1,0),colors.white),

        ("GRID",(0,0),(-1,-1),1,colors.black),

        ("BACKGROUND",(0,1),(-1,-1),colors.beige),

        ("ALIGN",(0,0),(-1,-1),"CENTER"),

        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ("BOTTOMPADDING",(0,0),(-1,0),10)

    ]))

    elementos.append(tabla)

    pdf.build(elementos)

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name="Reporte_Ventas.pdf",

        mimetype="application/pdf"

    )

@reportes_bp.route("/reportes/excel")
def reporte_excel():

    ventas = Venta.query.order_by(
        Venta.fecha.desc()
    ).all()

    libro = Workbook()

    hoja = libro.active

    hoja.title = "Reporte de Ventas"

    encabezados = [

        "ID",
        "Fecha",
        "Cliente",
        "Total"

    ]

    color = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for columna, texto in enumerate(encabezados, start=1):

        celda = hoja.cell(row=1, column=columna)

        celda.value = texto

        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = color

    fila = 2

    total = 0

    for venta in ventas:

        hoja.cell(fila, 1).value = venta.id

        hoja.cell(fila, 2).value = venta.fecha.strftime(
            "%d/%m/%Y %H:%M"
        )

        hoja.cell(fila, 3).value = venta.cliente.nombre

        hoja.cell(fila, 4).value = venta.total

        total += venta.total

        fila += 1

    hoja.cell(fila + 1, 3).value = "TOTAL"

    hoja.cell(fila + 1, 3).font = Font(bold=True)

    hoja.cell(fila + 1, 4).value = total

    hoja.cell(fila + 1, 4).font = Font(bold=True)

    for columna in hoja.columns:

        longitud = 0

        letra = get_column_letter(
            columna[0].column
        )

        for celda in columna:

            try:

                if len(str(celda.value)) > longitud:

                    longitud = len(str(celda.value))

            except:

                pass

        hoja.column_dimensions[letra].width = longitud + 5

    archivo = io.BytesIO()

    libro.save(archivo)

    archivo.seek(0)

    return send_file(

        archivo,

        download_name="Reporte_Ventas.xlsx",

        as_attachment=True,

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

